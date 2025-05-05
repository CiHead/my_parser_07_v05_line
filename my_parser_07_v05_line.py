# my_parser_07_v05_line.py

# pip install requests beautifulsoup4 lxml

'''
Вариант с использованием словаря list_card_url = [] для
считывания с сайта информации о карточках товара.
Этот вариант менее эффективен, так как при больших объёмах данных
словарь использует значительный ресурс памяти.

В этом варианте словарь list_card_url = [] заменён процедурой def get_url(),
в которой применяется инструкция не return, а yield card_url.

Добавлены строки пагинации для подсчёта количества страниц.

Добавлены строки для экономии памяти при парсинге - спарсенные карточки товаров
накапливаются не в списке, а в последовательно файле «data_XlsxWriter_line.xlsx».



П А Р С И Н Г

При работе с HTML и XML в Python часто возникает необходимость в парсинге, и для этого 
доступны различные парсеры, такие как `lxml` и встроенный в Python `html.parser`. 
Выбор между ними зависит от нескольких факторов, включая требования к 
производительности, функциональности и удобству использования.

lxml

`lxml` — это мощная библиотека для парсинга XML и HTML, которая предоставляет 
расширенные возможности и высокую производительность. 
Вот когда имеет смысл использовать `lxml`:

1. Производительность
`lxml` обычно быстрее, чем встроенный `html.parser`,  особенно при работе с большими 
объемами данных. Если вам важна скорость обработки, `lxml` может быть лучшим выбором.

2. Поддержка XML
`lxml` поддерживает как HTML, так и XML. 
Если вам нужно работать с обоими форматами или с более сложными 
XML-структурами, `lxml` предоставляет больше возможностей.

3. XPath и XSLT
`lxml` поддерживает XPath и XSLT, что делает его мощным 
инструментом для сложного поиска и трансформации документов.

4. Обработка некорректных HTML
Хотя `html.parser` также может обрабатывать некорректный HTML, 
`lxml` часто делает это более эффективно и корректно.

5. Расширенные возможности
`lxml` предлагает множество дополнительных функций, таких как работа с 
пространствами имен, поддержка DTD и схем валидации.

html.parser

`html.parser` — это стандартный парсер HTML, встроенный в Python. 
Его использование может быть предпочтительным в следующих случаях:

1. Простота использования
Поскольку это встроенный модуль, вам не нужно устанавливать дополнительные зависимости. 
Это может быть удобно для простых задач или скриптов, которые должны 
легко запускаться в разных средах.

2. Легкость
Для небольших проектов или простых задач, где высокая производительность не 
является критичной, `html.parser` может быть вполне достаточным.

3. Совместимость
`html.parser` доступен из коробки в Python, что делает его привлекательным 
для приложений, где установка внешних библиотек затруднительна.

4. Обработка некорректных HTML
`html.parser` способен обрабатывать некорректный HTML, что может быть 
полезно для веб-скрейпинга.

Итог

- Используйте `lxml`, если вы работаете с большими объемами данных, нуждаетесь в 
поддержке XML, XPath, XSLT, или если вам требуется высокая производительность и 
расширенные возможности.
- Используйте `html.parser`, если вы хотите избежать установки дополнительных 
библиотек, работаете с простыми задачами, или если вам нужна простота и совместимость.

Оба парсера имеют свои преимущества и недостатки, и выбор между ними должен 
основываться на специфических требованиях вашего проекта.

'''

# Импорт необходимых библиотек
import requests # Для отправки HTTP-запросов к сайту
from bs4 import BeautifulSoup # Для парсинга HTML-страниц
import pprint # Для красивого вывода данных в консоль
from time import sleep # Для добавления задержек между запросами
import pandas as pd # Для работы с данными в табличном формате
import os # Для работы с файловой системой
from openpyxl import Workbook, load_workbook # Для работы с Excel-файлами

# Базовый домен сайта (используется для формирования полных URL)
domen = 'https://scrapingclub.com'
# Стартовая URL-страница для парсинга
url = 'https://scrapingclub.com/exercise/list_basic/?page=1'

# list_card_url = []

headers = {
    "Accept": "*/*",
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Cache-Control": "no-cache",
    "Connection": "keep-alive",
    "Cookie": "sessionid=example_session_id; csrftoken=example_csrf_token",
    "DNT": "1", # Do Not Track Request Header
    "Host": "example.com",
    "Pragma": "no-cache",
    "Referer": "https://example.com/",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "same-origin",
    "Sec-Fetch-User": "?1",
    "TE": "Trailers",
    "Upgrade-Insecure-Requests": "1",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "X-Requested-With": "XMLHttpRequest"
}

# Заголовки HTTP-запроса (имитируем браузер)
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3"
}

# Путь к файлу Excel для сохранения результатов для построчного последовательного заполнения спарсенными карточками товаров
output_file = r'D:\PROJECTS\PROJECT_20220926_SKILL_FACTORY\IDE\WILDBERRIES\data_XlsxWriter_line.xlsx'

# Проверка существования файла и его создание/загрузка: ...
# ... если файл есть — загружаем, если нет — создаём
if not os.path.exists(output_file):
    # Если файл не существует - создаем новый
    wb = Workbook()  # Создаем новую книгу Excel
    ws = wb.active  # Получаем активный лист
    ws.title = "Sheet1"  # Переименовываем лист
    # Добавляем заголовки столбцов
    ws.append(['name', 'price', 'text', 'url_img']) # Заголовки
    wb.save(output_file)  # Сохраняем файл
    print(f"{os.path.basename(output_file)} создан")
else:
    # Если файл существует - загружаем его
    wb = load_workbook(output_file)
    ws = wb.active
    print(f"{os.path.basename(output_file)} создан повторно")

# Отправляем GET-запрос к стартовой странице
response = requests.get(url, headers = headers)

# Создаем объект BeautifulSoup для парсинга HTML
# Здесь использую парсер lxml, но в некоторых случаях можно html.parser вместо lxml
soup = BeautifulSoup(response.text, 'lxml')

# Находим блок пагинации
pagination = soup.find('nav', class_='pagination')
# Получаем все элементы пагинации (теги <a>)...
# ... pagination.find_all('a') возвращает список ссылок на страницы: 1 2 3 ... Next.
pages = pagination.find_all('a')
# Последний элемент перед "Next" — это номер последней страницы или иначе ...
# ... предпоследний элемент (pages[-2]) содержит номер последней страницы и ...
# ... потом преобразуем его .text в int.
last_page_number = int(pages[-2].text)
print("\nКоличество страниц:", last_page_number,'\n')

def get_url():
    """Генератор, который возвращает URL-адреса карточек товаров со всех страниц"""
    """Функция с yield, которая возвращает URL по одному - это позволяет ..."""
    """... обрабатывать страницы последовательно без загрузки всех URL в память"""

    for count in range(1, last_page_number):

        # Формируем URL для каждой страницы
        url = f'https://scrapingclub.com/exercise/list_basic/?page={count}'

        # Отправляем запрос к странице
        response = requests.get(url, headers = headers)
#        print('\nresponse:\n', response)
#        print('\nresponse.test:\n', response.text)
#        print('\nresponse.json:\n', response.json)

        # Парсим HTML страницы
        soup = BeautifulSoup(response.text, 'lxml') # в некоторых случаях можно html.parser вместо lxml
#        print('\nsoup:\n', soup)

        data = soup.find_all('div')
#        print('\ndata div:\n', data)

        # Находим все карточки товаров на странице
        data = soup.find_all('div', class_ = 'w-full rounded border')

        # Для каждой карточки извлекаем информацию
        for i in data:
#            print('\ndata div class:\n', i)
            # Извлекаем название товара (удаляем лишние пробелы и переносы строк)
            name = i.find('h4').text.replace(' ', '').replace('\n', '')
#            print('\nname =', name)
            # Извлекаем цену товара
            price = i.find('h5').text
#            print('price =', price)
            # Формируем полный URL изображения (добавляем домен)
            url_img = domen + i.find('img', class_ = 'card-img-top img-fluid').get('src')
#            print('url_img =', url_img)
            # Формируем полный URL страницы товара
            card_url = domen + i.find('a').get('href')
#            print('card_url =', card_url)
            # Возвращаем URL страницы товара (генератор)
            yield card_url

# Список (словарь) для хранения данных обо всех товарах по каждой карточке товара
data_list = []

# Итерируем по всем URL карточек товаров
for card_url in get_url():
    # Отправляем запрос к странице товара
    response = requests.get(card_url, headers=headers)
    # Парсим HTML страницы товара
    soup = BeautifulSoup(response.text, 'lxml')

    # Находим основной блок с информацией о товаре
    data = soup.find('div', class_='my-8 w-full rounded border')
#    print('data =', data)
    
    if data: # Если блок найден
        try:
            # Извлекаем название товара
            name = data.find('h3', class_='card-title').text.strip()
            # Извлекаем цену товара
            price = data.find('h4', class_='my-4 card-price').text.strip()
            # Извлекаем описание товара
            text = data.find('p', class_='card-description').text.strip()
            # Формируем полный URL изображения
            url_img = domen + data.find('img').get('src')

            # Добавляем данные о товаре в список - пополнение словаря
            data_list.append({
                'name': name,
                'price': price,
                'text': text,
                'url_img': url_img
            })

            # Добавляем данные в Excel-файл (построчно)
            ws.append([name, price, text, url_img])
            # Сохраняем изменения в файле
            wb.save(output_file)

        except AttributeError as e:
            # Если какой-то элемент не найден
            print(f"Некоторые данные не найдены на странице: {card_url}")
            print(f"Ошибка: {e}")
    else:
        print(f"Блок данных не найден на странице: {card_url}")

# Создаем DataFrame из собранных данных
df = pd.DataFrame(data_list, columns=['name', 'price', 'text', 'url_img'])

# Выводим DataFrame в консоль
print(df)

# Сохраняем DataFrame в отдельный Excel-файл
df.to_excel(r'D:\PROJECTS\PROJECT_20220926_SKILL_FACTORY\IDE\WILDBERRIES\data_XlsxWriter.xlsx', index=False)

# Завершаем работу программы
quit()
